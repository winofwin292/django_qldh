import traceback
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, Http404
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import json
import openpyxl
import os
from datetime import datetime, timedelta
from .cal_setup import get_calendar_service

from .models import CustomUser, TrinhDoHocVan, PhongHoc, LopHoc, MonHoc, NamHoc, HocKy, GiaoVien, HocSinh, GiangDay
from .forms import AddTeacherForm, EditTeacherForm, AddClassroomForm, EditClassroomForm, AddStudentForm, EditStudentForm


def auto_ids(model_name, keyword, column_id):
    try:
        obj = model_name.objects.latest(column_id)
        no = str(obj.get_id())
        id = int(no[-6:]) + 1
    except :
        id = 1

    return "{}{:06d}".format(keyword, id)


def auto_email(str_id):
    return str_id + '@gmail.com'


def download_excel(request, file_name):
    path = 'excel/' + file_name
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(),
                                    content_type="application/vnd.ms-excel, "
                                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404


def admin_home(request):
    so_giao_vien = GiaoVien.objects.all().count()
    so_hoc_sinh = HocSinh.objects.all().count()
    so_lop_hoc = LopHoc.objects.all().count()
    so_mon_hoc = MonHoc.objects.all().count()

    context = {
        "so_giao_vien": so_giao_vien,
        "so_hoc_sinh": so_hoc_sinh,
        "so_lop_hoc": so_lop_hoc,
        "so_mon_hoc": so_mon_hoc,
    }
    return render(request, 'admin_templates/home_content.html', context)


@csrf_exempt
def check_email_exist(request):
    email = request.POST.get("email")
    user_obj = CustomUser.objects.filter(email=email).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


@csrf_exempt
def check_username_exist(request):
    username = request.POST.get("username")
    user_obj = CustomUser.objects.filter(username=username).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


def manage_teacher(request):
    teachers = GiaoVien.objects.all()
    mh = MonHoc.objects.all()
    form = AddTeacherForm()
    context = {
        "Teachers": teachers,
        "mh": mh,
        "form": form,
    }
    return render(request, 'admin_templates/manage_teacher_template.html', context)


@csrf_exempt
def admin_get_teacher(request):
    if request.accepts("application/json") and request.method == "POST":
        ma_mon = request.POST.get('subject')
        try:
            if ma_mon == 'all':
                giao_vien = GiaoVien.objects.all()
            else:
                subject_model = MonHoc.objects.get(ma_mon=ma_mon)
                giao_vien = GiaoVien.objects.filter(day_mon_id=subject_model)
            list_data = []
            for tc in giao_vien:
                small_data = {"magv": tc.magv.username, "hoten": tc.magv.last_name + " " + tc.magv.first_name,
                              "email": tc.magv.email, "que_quan": tc.que_quan, "so_dien_thoai": tc.so_dien_thoai,
                              "gioi_tinh": tc.gioi_tinh, "ngay_sinh": tc.ngay_sinh.strftime("%d/%m/%Y"),
                              "day_mon": tc.day_mon.ten_mon, "trinh_do": tc.trinh_do.mo_ta_trinh_do,
                              "profile_pic": str(tc.profile_pic)}
                list_data.append(small_data)
            return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
        except:
            return JsonResponse({"error": ""}, status=400)
    return JsonResponse({"error": ""}, status=400)


def add_teacher_from_xls(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_teacher')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            try:
                user = CustomUser.objects.create_user(username=auto_ids(GiaoVien, 'GV', 'magv_id'),
                                                      password=auto_ids(GiaoVien, 'GV', 'magv_id'),
                                                      email=auto_email(auto_ids(GiaoVien, 'GV', 'magv_id')),
                                                      first_name=str(row[1].value), last_name=str(row[0].value),
                                                      user_type=2)
                user.giaovien.que_quan = str(row[2].value)
                user.giaovien.so_dien_thoai = str(row[3].value)
                user.giaovien.gioi_tinh = str(row[4].value)
                user.giaovien.ngay_sinh = row[5].value
                td = TrinhDoHocVan.objects.get(mo_ta_trinh_do=str(row[6].value))
                user.giaovien.trinh_do = td
                mon = MonHoc.objects.get(ten_mon=str(row[7].value))
                user.giaovien.day_mon = mon
                user.save()
            except:
                messages.error(request, "Thêm giáo viên mới không thành công!, Lỗi ở dòng: " + str(row[0].row))
                return redirect('manage_teacher')
        messages.success(request, "Thêm giáo viên thành công!!!")
        return redirect('manage_teacher')


def add_teacher_save(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_teacher')
    else:
        form = AddTeacherForm(request.POST, request.FILES)

        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            que_quan = form.cleaned_data['que_quan']
            ngay_sinh = form.cleaned_data['ngay_sinh']
            so_dien_thoai = form.cleaned_data['so_dien_thoai']
            gioi_tinh = form.cleaned_data['gioi_tinh']
            trinh_do = form.cleaned_data['trinh_do']
            day_mon = form.cleaned_data['day_mon']
            try:
                user = CustomUser.objects.create_user(username=auto_ids(GiaoVien, 'GV', column_id='magv_id'),
                                                      password=auto_ids(GiaoVien, 'GV', column_id='magv_id'),
                                                      email=auto_email(auto_ids(GiaoVien, 'GV', column_id='magv_id')),
                                                      first_name=first_name, last_name=last_name, user_type=2)
                user.giaovien.que_quan = que_quan
                user.giaovien.so_dien_thoai = so_dien_thoai

                td = TrinhDoHocVan.objects.get(trinh_do=trinh_do)
                user.giaovien.trinh_do = td

                mon = MonHoc.objects.get(ma_mon=day_mon)
                user.giaovien.day_mon = mon

                user.giaovien.gioi_tinh = gioi_tinh
                user.giaovien.ngay_sinh = ngay_sinh
                user.save()
                messages.success(request, "Tạo giáo viên thành công!")
                return redirect('manage_teacher')
            except:
                messages.error(request, "Tạo không thành công!!")
                return redirect('manage_teacher')
        else:
            return redirect('manage_teacher')


def edit_teacher(request, magv):
    request.session['magv'] = magv
    gv = GiaoVien.objects.get(magv=magv)
    form = EditTeacherForm()

    form.fields['username'].initial = gv.magv.username
    form.fields['email'].initial = gv.magv.email
    form.fields['first_name'].initial = gv.magv.first_name
    form.fields['last_name'].initial = gv.magv.last_name
    form.fields['que_quan'].initial = gv.que_quan
    form.fields['ngay_sinh'].initial = gv.ngay_sinh
    form.fields['so_dien_thoai'].initial = gv.so_dien_thoai
    form.fields['gioi_tinh'].initial = gv.gioi_tinh
    form.fields['day_mon'].initial = gv.day_mon.ma_mon
    form.fields['trinh_do'].initial = gv.trinh_do.trinh_do

    context = {
        "magv": magv,
        "form": form
    }

    return render(request, 'admin_templates/edit_teacher_templates.html', context)


def edit_teacher_save(request):
    if request.method != "POST":
        return HttpResponse("Lỗi phương thức")
    else:
        magv = request.session.get('magv')
        if magv == None:
            return redirect('/qldh/quan_ly_giao_vien/')

        form = EditTeacherForm(request.POST, request.FILES)
        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            que_quan = form.cleaned_data['que_quan']
            so_dien_thoai = form.cleaned_data['so_dien_thoai']
            ngay_sinh = form.cleaned_data['ngay_sinh']
            gioi_tinh = form.cleaned_data['gioi_tinh']
            trinh_do = form.cleaned_data['trinh_do']
            day_mon = form.cleaned_data['day_mon']

            try:
                user = CustomUser.objects.get(username=magv)
                user.first_name = first_name
                user.last_name = last_name
                user.save()

                gv = GiaoVien.objects.get(magv=magv)
                gv.que_quan = que_quan
                gv.so_dien_thoai = so_dien_thoai
                gv.gioi_tinh = gioi_tinh
                gv.ngay_sinh = ngay_sinh
                td = TrinhDoHocVan.objects.get(trinh_do=trinh_do)
                gv.trinh_do = td
                mon = MonHoc.objects.get(ma_mon=day_mon)
                gv.day_mon = mon
                gv.save()
                del request.session['magv']
                messages.success(request, "Chỉnh sửa thông tin giáo viên thành công!")
                return redirect('/qldh/chinh_sua_giao_vien/' + magv)
            except:
                messages.success(request, "Lỗi khi cập nhât.")
                return redirect('/qldh/chinh_sua_giao_vien/' + magv)
        else:
            return redirect('/qldh/chinh_sua_giao_vien/' + magv)


def delete_teacher(request, magv):
    user = CustomUser.objects.get(username=magv)
    try:
        user.delete()
        messages.success(request, "Xóa giáo viên thành công.")
        return redirect('manage_teacher')
    except:
        messages.error(request, "Lỗi khi xóa giáo viên.")
        return redirect('manage_teacher')


def manage_study_room(request):
    study_rooms = PhongHoc.objects.all()
    context = {
        "study_rooms": study_rooms
    }
    return render(request, 'admin_templates/manage_study_room_template.html', context)


def add_study_room_from_xls(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_study_room')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            try:
                ph = PhongHoc.objects.create(ma_phong=auto_ids(model_name=PhongHoc, keyword="PH", column_id='ma_phong'),
                                             ten_phong=str(row[0].value),
                                             so_luong_cho_ngoi=row[1].value)
                ph.save()
            except:
                messages.error(request, "Thêm phòng học mới không thành công!, Lỗi ở dòng: " + str(row[0].row))
                return redirect('manage_study_room')

        messages.success(request, "Thêm phòng học thành công!!!")
        return redirect('manage_study_room')


def add_study_room_save(request):
    if request.method != "POST":
        return HttpResponse("Lỗi phương thức")
    else:
        ten_phong = request.POST.get('ten_phong')
        so_luong_cho_ngoi = request.POST.get('so_luong_cho_ngoi')
        try:
            ph = PhongHoc.objects.create(ma_phong=auto_ids(model_name=PhongHoc, keyword="PH", column_id='ma_phong'), ten_phong=ten_phong,
                                         so_luong_cho_ngoi=so_luong_cho_ngoi)
            ph.save()
            messages.success(request, "Tạo phòng học thành công!!!")
            return redirect('manage_study_room')
        except:
            messages.error(request, "Tạo không thành công!")
            return redirect('manage_study_room')


def edit_study_room(request, ma_phong):
    ph = PhongHoc.objects.get(ma_phong=ma_phong)
    context = {
        "phong_hoc": ph
    }
    return render(request, 'admin_templates/edit_study_room_template.html', context)


def edit_study_room_save(request):
    if request.method != "POST":
        HttpResponse("Lỗi phương thức")
    else:
        ma_phong = request.POST.get('ma_phong')
        ten_phong = request.POST.get('ten_phong')
        so_luong_cho_ngoi = request.POST.get('so_luong_cho_ngoi')

        try:
            phong_hoc = PhongHoc.objects.get(ma_phong=ma_phong)
            phong_hoc.ten_phong = ten_phong
            phong_hoc.so_luong_cho_ngoi = so_luong_cho_ngoi
            phong_hoc.save()

            messages.success(request, "Chỉnh sửa thành công.")
            return redirect('/qldh/chinh_sua_phong_hoc/' + ma_phong)

        except:
            messages.error(request, "Chỉnh sửa không thành công.")
            return redirect('/qldh/chinh_sua_phong_hoc/' + ma_phong)


def delete_study_room(request, ma_phong):
    phong_hoc = PhongHoc.objects.get(ma_phong=ma_phong)
    try:
        phong_hoc.delete()
        messages.success(request, "Xóa phòng học thành công.")
        return redirect('manage_study_room')
    except:
        messages.error(request, "Lỗi khi xóa phòng học.")
        return redirect('manage_study_room')


def manage_classroom(request):
    lh = LopHoc.objects.all()
    form = AddClassroomForm()
    grade_list = ('Khối 10', 'Khối 11', 'Khối 12')
    context = {
        "lop_hoc": lh,
        "khoi": grade_list,
        "form": form,
    }
    return render(request, 'admin_templates/manage_classroom_template.html', context)


@csrf_exempt
def admin_get_classroom(request):
    if request.accepts("application/json") and request.method == "POST":
        khoi = request.POST.get('khoi')
        try:
            if khoi == 'all':
                lop_hoc = LopHoc.objects.all()
            else:
                lop_hoc = LopHoc.objects.filter(khoi=khoi)
            list_data = []
            for k in lop_hoc:
                small_data = {"ma_lop": k.ma_lop, "ten_lop": k.ten_lop, "phong": k.phong.ten_phong,
                              "giao_vien_chu_nhiem": k.giao_vien_chu_nhiem.magv.last_name + " " + k.giao_vien_chu_nhiem.magv.first_name,
                              "si_so": k.si_so, "khoi": k.khoi, "nam_hoc": k.nam_hoc.mo_ta}
                list_data.append(small_data)
            return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
        except:
            return JsonResponse({"error": ""}, status=400)
    return JsonResponse({"error": ""}, status=400)


def add_classroom_from_xls(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_classroom')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            try:
                nh = NamHoc.objects.get(mo_ta=str(row[1].value))
                ph = PhongHoc.objects.get(ten_phong=str(row[3].value))
                gvcn = GiaoVien.objects.get(magv=str(row[4].value))
                lh = LopHoc.objects.create(ma_lop=auto_ids(LopHoc, "LH", column_id='ma_lop'), ten_lop=str(row[0].value), nam_hoc=nh,
                                           khoi=str(row[2].value),
                                           phong=ph, si_so=0, giao_vien_chu_nhiem=gvcn)
                lh.save()
            except Exception as e:
                traceback.print_exc()
                messages.error(request, "Thêm lớp học không thành công!, Lỗi ở dòng: " + str(row[0].row))
                return redirect('manage_classroom')

        messages.success(request, "Thêm lớp học thành công!!!")
        return redirect('manage_classroom')


def add_classroom_save(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_classroom')
    else:
        form = AddClassroomForm(request.POST)

        if form.is_valid():
            ten_lop = form.cleaned_data['ten_lop']
            nam_hoc = NamHoc.objects.get(hien_tai=True)
            khoi = form.cleaned_data['khoi']
            phong = form.cleaned_data['phong']
            giao_vien = form.cleaned_data['giao_vien_chu_nhiem']

            try:
                lh = LopHoc.objects.create(ma_lop=auto_ids(LopHoc, "LH", column_id='ma_lop'), ten_lop=ten_lop,
                                           nam_hoc=nam_hoc, khoi=khoi, phong_id=phong, si_so=0,
                                           giao_vien_chu_nhiem_id=giao_vien)
                lh.save()
                messages.success(request, "Tạo lớp học thành công!")
                return redirect('manage_classroom')
            except:
                messages.error(request, "Tạo không thành công!!")
                return redirect('manage_classroom')
        else:
            return redirect('manage_classroom')


def edit_classroom(request, ma_lop):
    request.session['malop'] = ma_lop
    lh = LopHoc.objects.get(ma_lop=ma_lop)
    form = EditClassroomForm()

    form.fields["ma_lop"].initial = lh.ma_lop
    form.fields["ten_lop"].initial = lh.ten_lop
    form.fields["khoi"].initial = lh.khoi
    form.fields["phong"].initial = lh.phong.ma_phong
    form.fields["giao_vien_chu_nhiem"].initial = lh.giao_vien_chu_nhiem.magv
    context = {
        "form": form,
        "ma_lop": ma_lop
    }
    return render(request, 'admin_templates/edit_classroom_template.html', context)


def edit_classroom_save(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('edit_classroom')
    else:
        ma_lop = request.session.get('malop')
        if ma_lop == None:
            return redirect('/qldh/quan_ly_lop_hoc')
        form = EditClassroomForm(request.POST)

        if form.is_valid():
            ten_lop = form.cleaned_data['ten_lop']
            nam_hoc = form.cleaned_data['nam_hoc']
            khoi = form.cleaned_data['khoi']
            phong = form.cleaned_data['phong']
            giao_vien = form.cleaned_data['giao_vien_chu_nhiem']

            try:
                lh = LopHoc.objects.get(ma_lop=ma_lop)
                lh.ten_lop = ten_lop
                lh.nam_hoc = NamHoc.objects.get(nam_hoc=nam_hoc)
                lh.khoi = khoi
                lh.phong = PhongHoc.objects.get(ma_phong=phong)
                lh.giao_vien_chu_nhiem = GiaoVien.objects.get(magv=giao_vien)
                lh.save()
                del request.session['malop']
                messages.success(request, "Chỉnh sửa lớp học thành công!")
                return redirect('/qldh/chinh_sua_lop_hoc/' + ma_lop)
            except Exception:
                traceback.print_exc()
                messages.error(request, "Chỉnh sửa không thành công!!")
                return redirect('/qldh/chinh_sua_lop_hoc/' + ma_lop)
        else:
            messages.error(request, "Chỉnh sửa không thành công!!")
            return redirect('/qldh/chinh_sua_lop_hoc/' + ma_lop)


def delete_classroom(request, ma_lop):
    lh = LopHoc.objects.get(ma_lop=ma_lop)
    try:
        lh.delete()
        messages.success(request, "Xóa lớp học thành công!")
        return redirect('manage_classroom')
    except:
        messages.error(request, "Xóa không thành công!!")
        return redirect('manage_classroom')


def manage_student(request):
    students = HocSinh.objects.all()
    form = AddStudentForm()
    lh = LopHoc.objects.all()
    context = {
        "students": students,
        "lh": lh,
        "form": form
    }
    return render(request, 'admin_templates/manage_student_template.html', context)


@csrf_exempt
def admin_get_student(request):
    if request.accepts("application/json") and request.method == "POST":
        ma_lop = request.POST.get('classroom')
        try:
            if ma_lop == 'all':
                hoc_sinh = HocSinh.objects.all()
            else:
                classroom_model = LopHoc.objects.get(ma_lop=ma_lop)
                hoc_sinh = HocSinh.objects.filter(lop=classroom_model)
            list_data = []
            for hs in hoc_sinh:
                small_data = {"mahs": hs.mahs.username, "hoten": hs.mahs.last_name + " " + hs.mahs.first_name,
                              "email": hs.mahs.email, "que_quan": hs.que_quan,
                              "ngay_sinh": hs.ngay_sinh.strftime("%d/%m/%Y"), "gioi_tinh": hs.gioi_tinh,
                              "dan_toc": hs.dan_toc, "lop": hs.lop.ten_lop, "profile_pic": str(hs.profile_pic)}
                list_data.append(small_data)
            return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
        except:
            return JsonResponse({"error": ""}, status=400)
    return JsonResponse({"error": ""}, status=400)


def add_student_from_xls(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_student')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            try:
                user = CustomUser.objects.create_user(username=auto_ids(HocSinh, 'HS', column_id='mahs_id'),
                                                      password=auto_ids(HocSinh, 'HS', column_id='mahs_id'),
                                                      email=auto_email(auto_ids(HocSinh, 'HS', column_id='mahs_id')),
                                                      first_name=str(row[1].value), last_name=str(row[0].value),
                                                      user_type=3)
                user.hocsinh.que_quan = str(row[2].value)
                user.hocsinh.gioi_tinh = str(row[3].value)
                user.hocsinh.ngay_sinh = row[4].value
                user.hocsinh.dan_toc = str(row[5].value)
                lop = LopHoc.objects.get(ten_lop=str(row[6].value))
                user.hocsinh.lop = lop
                lop.si_so += 1
                lop.save()
                user.save()
            except Exception:
                traceback.print_exc()
                messages.error(request, "Thêm học sinh mới không thành công!, Lỗi ở dòng: " + str(row[0].row))
                return redirect('manage_student')
        messages.success(request, "Thêm học sinh thành công!!!")
        return redirect('manage_student')


def add_student_save(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_student')
    else:
        form = AddStudentForm(request.POST, request.FILES)

        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            que_quan = form.cleaned_data['que_quan']
            ngay_sinh = form.cleaned_data['ngay_sinh']
            gioi_tinh = form.cleaned_data['gioi_tinh']
            lop = form.cleaned_data['lop']
            dan_toc = form.cleaned_data['dan_toc']

            try:
                user = CustomUser.objects.create_user(username=auto_ids(HocSinh, 'HS', column_id='mahs_id'),
                                                      password=auto_ids(HocSinh, 'HS', column_id='mahs_id'),
                                                      email=auto_email(auto_ids(HocSinh, 'HS', column_id='mahs_id')),
                                                      first_name=first_name, last_name=last_name, user_type=3)
                user.hocsinh.que_quan = que_quan
                user.hocsinh.dan_toc = dan_toc
                user.hocsinh.gioi_tinh = gioi_tinh
                user.hocsinh.ngay_sinh = ngay_sinh

                lop = LopHoc.objects.get(ma_lop=lop)
                user.hocsinh.lop = lop
                lop.si_so += 1
                lop.save()

                user.save()
                messages.success(request, "Tạo học sinh thành công!")
                return redirect('manage_student')
            except Exception:
                traceback.print_exc()
                messages.error(request, "Tạo không thành công!!")
                return redirect('manage_student')
        else:
            return redirect('manage_student')


def edit_student(request, mahs):
    request.session['mahs'] = mahs
    hs = HocSinh.objects.get(mahs=mahs)
    form = EditStudentForm()

    form.fields['username'].initial = hs.mahs.username
    form.fields['email'].initial = hs.mahs.email
    form.fields['first_name'].initial = hs.mahs.first_name
    form.fields['last_name'].initial = hs.mahs.last_name
    form.fields['que_quan'].initial = hs.que_quan
    form.fields['dan_toc'].initial = hs.dan_toc
    form.fields['gioi_tinh'].initial = hs.gioi_tinh
    form.fields['ngay_sinh'].initial = hs.ngay_sinh
    form.fields['lop'].initial = hs.lop.ma_lop

    context = {
        "mahs": mahs,
        "form": form
    }
    return render(request, 'admin_templates/edit_student_template.html', context)


def edit_student_save(request):
    if request.method != "POST":
        return HttpResponse("Lỗi phương thức")
    else:
        mahs = request.session.get('mahs')
        if mahs == None:
            return redirect('manage_student')

        form = EditStudentForm(request.POST, request.FILES)
        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            email = form.cleaned_data['email']
            que_quan = form.cleaned_data['que_quan']
            dan_toc = form.cleaned_data['dan_toc']
            gioi_tinh = form.cleaned_data['gioi_tinh']
            ngay_sinh = form.cleaned_data['ngay_sinh']
            lop = form.cleaned_data['lop']

            if len(request.FILES) != 0:
                profile_pic = request.FILES['profile_pic']
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                profile_pic_url = fs.url(filename)
            else:
                profile_pic_url = None

            try:
                user = CustomUser.objects.get(username=mahs)
                user.first_name = first_name
                user.last_name = last_name
                user.email = email
                user.save()

                hs = HocSinh.objects.get(mahs=mahs)
                hs.que_quan = que_quan
                hs.dan_toc = dan_toc
                hs.gioi_tinh = gioi_tinh
                hs.ngay_sinh = ngay_sinh
                lop_hoc = LopHoc.objects.get(ma_lop=lop)
                hs.lop = lop_hoc

                if profile_pic_url != None:
                    hs.profile_pic = profile_pic_url
                hs.save()

                del request.session['mahs']

                messages.success(request, "Chỉnh sửa thông tin học sinh thành công!")
                return redirect('manage_student')
            except:
                messages.success(request, "Lỗi khi cập nhât.")
                return redirect('manage_student')
        else:
            return redirect('manage_student')


def delete_student(request, mahs):
    hs = HocSinh.objects.get(mahs=mahs)
    user = CustomUser.objects.get(username=mahs)
    try:
        hs.lop.si_so -= 1
        hs.lop.save()
        user.delete()
        messages.success(request, "Xóa học sinh thành công.")
        return redirect('manage_student')
    except:
        messages.error(request, "Lỗi khi xóa học sinh.")
        return redirect('manage_student')


@csrf_exempt
def get_teacher(request):
    if request.accepts("application/json") and request.method == "POST":
        ma_mon = request.POST.get('subject')
        try:
            subject_model = MonHoc.objects.get(ma_mon=ma_mon)
            giao_vien = GiaoVien.objects.filter(day_mon_id=subject_model)
            list_data = []
            for gv in giao_vien:
                small_data = {"magv": gv.magv.username, "ho_ten": str(gv.magv.last_name + " " + gv.magv.first_name)}
                list_data.append(small_data)
            return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
        except:
            return JsonResponse({"error": ""}, status=400)
    return JsonResponse({"error": ""}, status=400)


@csrf_exempt
def get_tuition(request):
    if request.accepts("application/json") and request.method == "POST":
        giao_vien = request.POST.get('teacher')
        try:
            gv = GiaoVien.objects.get(magv=giao_vien)
            nam_hoc = NamHoc.objects.get(hien_tai=True)
            hoc_ky = HocKy.objects.get(hien_tai=True)
            filter_gd = GiangDay.objects.filter(magv=gv, nam_hoc=nam_hoc, hoc_ky=hoc_ky)
            list_gd = []
            for item in filter_gd:
                small_data = {"id": item.id, "thu": item.thu, "ma_lop": item.ma_lop.ten_lop, "1": item.tiet_1, "2": item.tiet_2, "3": item.tiet_3
                              , "4": item.tiet_4, "5": item.tiet_5, "6": item.tiet_6, "7": item.tiet_7, "8": item.tiet_8
                              , "9": item.tiet_9}
                list_gd.append(small_data)
            return JsonResponse(json.dumps(list_gd), content_type="application/json", safe=False)
        except Exception:
            traceback.print_exc()
            return JsonResponse({"error": "try"}, status=400)
    return JsonResponse({"error": "if"}, status=400)


def them_moi_giang_day(request):
    if request.accepts("application/json") and request.method == "POST":
        nam_hoc = NamHoc.objects.get(hien_tai=True)
        hoc_ky = HocKy.objects.get(hien_tai=True)
        try:
            giao_vien = request.POST.get('teacher')
            ma_lop = request.POST.get('ma_lop')
            thu = request.POST.get('thu')
            tiet = request.POST.get('tiet')
            gv = GiaoVien.objects.get(magv=giao_vien)
            lop = LopHoc.objects.get(ma_lop=ma_lop)

            if GiangDay.objects.filter(magv=gv, ma_lop=lop, nam_hoc=nam_hoc, hoc_ky=hoc_ky).count() > gv.day_mon.tiet_tuan:
                return JsonResponse({"error": "Lỗi: số lượng tiết trên tuần quá giới hạn"}, status=400)

            gd = GiangDay.objects.create(magv=gv, ma_lop=lop, thu=thu, nam_hoc=nam_hoc, hoc_ky=hoc_ky)
            hour_start = 0
            minute_start = 0
            match tiet:
                case '1':
                    gd.tiet_1 = True
                    hour_start = 7
                    minute_start = 0
                case '2':
                    gd.tiet_2 = True
                    hour_start = 7
                    minute_start = 45
                case '3':
                    gd.tiet_3 = True
                    hour_start = 8
                    minute_start = 50
                case '4':
                    gd.tiet_4 = True
                    hour_start = 9
                    minute_start = 35
                case '5':
                    gd.tiet_5 = True
                    hour_start = 10
                    minute_start = 30
                case '6':
                    gd.tiet_6 = True
                    hour_start = 13
                    minute_start = 30
                case '7':
                    gd.tiet_7 = True
                    hour_start = 14
                    minute_start = 15
                case '8':
                    gd.tiet_8 = True
                    hour_start = 15
                    minute_start = 20
                case '9':
                    gd.tiet_9 = True
                    hour_start = 16
                    minute_start = 10
            #Thêm mới sự kiện trên google cal

            service = get_calendar_service()
            now_datetime = datetime.now().date()
            now_dateofweek = datetime.now().isoweekday()
            start_time = datetime(now_datetime.year, now_datetime.month, now_datetime.day,
                                  hour_start, minute_start) + timedelta(days=((int(thu)-1)-now_dateofweek))
            end_time = start_time + timedelta(minutes=45)
            event_result = service.events().insert(calendarId='primary',
                                                   body={
                                                       "summary": lop.ten_lop,
                                                       "description": lop.ten_lop,
                                                       "start": {"dateTime": start_time.isoformat(),
                                                                 "timeZone": 'Asia/Ho_Chi_Minh'},
                                                       "end": {"dateTime": end_time.isoformat(),
                                                               "timeZone": 'Asia/Ho_Chi_Minh'},
                                                       'recurrence': [
                                                           'RRULE:FREQ=WEEKLY;COUNT=18'
                                                       ],
                                                       # 'attendees': [
                                                       #     {'email': 'winofwin292@gmail.com'},
                                                       # ],
                                                   }
                                                   ).execute()
            gd.cal_id = event_result['id']

            gd.save()
            return JsonResponse({"success": "Thêm mới thành công"}, content_type="application/json", safe=False)
        except Exception:
            traceback.print_exc()
            return JsonResponse({"error": "Lỗi: Thêm mới không thành công"}, status=400)
    return JsonResponse({"error": "Lỗi: Sai phương thức"}, status=400)


@csrf_exempt
def chinh_sua_lop_giang_day(request):
    if request.accepts("application/json") and request.method == "POST":
        try:
            id = request.POST.get('id')
            ma_lop = request.POST.get('ma_lop')
            lop = LopHoc.objects.get(ma_lop=ma_lop)

            giang_day = GiangDay.objects.get(id=id)
            giang_day.ma_lop = lop

            service = get_calendar_service()
            event = service.events().get(calendarId='primary', eventId=giang_day.cal_id).execute()
            event['summary'] = lop.ten_lop
            event["description"] = lop.ten_lop
            updated_event = service.events().update(calendarId='primary', eventId=event['id'], body=event).execute()
            giang_day.save()
            return JsonResponse({"success": "Chỉnh sửa thành công"}, content_type="application/json", safe=False)
        except Exception:
            traceback.print_exc()
            return JsonResponse({"error": "Lỗi: Chỉnh sửa không thành công"}, status=400)
    return JsonResponse({"error": "Lỗi: Sai phương thức"}, status=400)


@csrf_exempt
def xoa_giang_day(request):
    if request.accepts("application/json") and request.method == "POST":
        try:
            id = request.POST.get('id')
            giang_day = GiangDay.objects.get(id=id)

            service = get_calendar_service()
            service.events().delete(calendarId='primary', eventId=giang_day.cal_id).execute()

            giang_day.delete()
            return JsonResponse({"success": "Xóa thành công"}, content_type="application/json", safe=False)
        except Exception:
            traceback.print_exc()
            return JsonResponse({"error": "Lỗi: Xóa không thành công"}, status=400)
    return JsonResponse({"error": "Lỗi: Sai phương thức"}, status=400)


def manage_tuition(request):
    mon = MonHoc.objects.all()
    lop = LopHoc.objects.all()
    context = {
        "mon": mon,
        "lop": lop,
    }
    return render(request, 'admin_templates/manage_tuition_template.html', context)


def manage_subject(request):
    mh = MonHoc.objects.all()
    context = {
        "mh": mh,
    }
    return render(request, 'admin_templates/manage_subject_template.html', context)


@csrf_exempt
def kiem_tra_ma_mon(request):
    ma_mon = request.POST.get("ma_mon")
    mh = MonHoc.objects.filter(ma_mon=ma_mon).exists()
    if mh:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


@csrf_exempt
def kiem_tra_ten_mon(request):
    ten_mon = request.POST.get("ten_mon")
    mh = MonHoc.objects.filter(ten_mon=ten_mon).exists()
    if mh:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


def add_subject_from_xls(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('manage_subject')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            try:
                mh = MonHoc.objects.create(ma_mon=str(row[0].value), ten_mon=str(row[1].value))
                mh.save()
            except:
                messages.error(request, "Thêm môn mới không thành công!, Lỗi ở dòng: " + str(row[0].row))
                return redirect('manage_subject')

        messages.success(request, "Thêm môn học thành công!!!")
        return redirect('manage_subject')


def add_subject_save(request):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('add_subject')
    else:
        ma_mon = request.POST.get('ma_mon')
        ten_mon = request.POST.get('ten_mon')
        if (ma_mon != '') or (ten_mon != ''):
            try:
                mh = MonHoc.objects.create(ma_mon=ma_mon, ten_mon=ten_mon)
                mh.save()
                messages.success(request, "Thêm môn học thành công!!!")
                return redirect('manage_subject')
            except:
                messages.error(request, "Thêm môn mới không thành công!")
                return redirect('manage_subject')
        else:
            messages.error(request, "Thêm môn mới không thành công! Dữ liệu không được rỗng")
            return redirect('manage_subject')


def edit_subject(request, ma_mon):
    mh = MonHoc.objects.get(ma_mon=ma_mon)
    context = {
        "mh": mh,
    }
    return render(request, 'admin_templates/edit_subject_template.html', context)


def edit_subject_save(request, ma_mon):
    if request.method != "POST":
        messages.error(request, "Lỗi phương thức")
        return redirect('/qldh/chinh_sua_mon_hoc/' + ma_mon)
    else:
        ten_mon = request.POST.get('ten_mon')

        try:
            mon_hoc = MonHoc.objects.get(ma_mon=ma_mon)
            mon_hoc.ten_mon = ten_mon
            mon_hoc.save()

            messages.success(request, "Chỉnh sửa môn học thành công.")
            return redirect('manage_subject')

        except:
            messages.error(request, "Chỉnh sửa môn học không thành công.")
            return redirect('/qldh/chinh_sua_mon_hoc/' + ma_mon)


def delete_subject(request, ma_mon):
    mon_hoc = MonHoc.objects.get(ma_mon=ma_mon)
    try:
        mon_hoc.delete()
        messages.success(request, "Xóa môn học thành công.")
        return redirect('manage_subject')
    except:
        messages.error(request, "Lỗi khi xóa môn học.")
        return redirect('manage_subject')
